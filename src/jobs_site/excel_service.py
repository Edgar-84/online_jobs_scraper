import openpyxl

from .base_exceptions import ExcelServiceExceptions
from src.config.logger_config import logger
from src.utils.BaseUtils import BaseInformationService


class ExcelService:
    @staticmethod
    def create_excel(save_data: dict, name_save: str) -> bool | None:
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            head_data = None
            all_data_values = []

            for i in save_data:
                head_data = save_data[i].keys()
                logger.info(head_data)
                break

            for i in save_data:
                all_data_values.append(save_data[i].values())

            column = 1
            for key in head_data:
                sheet.cell(row=1, column=column).value = key
                column += 1

            column = 1
            row = 2
            logger.debug("Start save excel")

            for num, values in enumerate(all_data_values):
                for value in values:
                    if column == 5:
                        sheet.cell(row=row, column=column).value = ' , '.join(value)
                        column += 1
                        continue
                    sheet.cell(row=row, column=column).value = value
                    column += 1

                row += 1
                column = 1

            logger.debug('Final')
            workbook.save(name_save)
            return True

        except ExcelServiceExceptions as ex:
            error_message = BaseInformationService.get_error_message(ex)
            logger.error(error_message)
            return None
